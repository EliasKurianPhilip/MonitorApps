import psutil
import time
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

# Initialize workbook and worksheet
workbook_name = "data_usage.xlsx"

try:
    wb = load_workbook(workbook_name)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Usage"
    ws.append(["App Name", "Data Usage (MB)"])

def get_network_usage():
    app_data = defaultdict(lambda: [0, 0])  # Dictionary to hold bytes_sent and bytes_recv for each app

    for proc in psutil.process_iter(['pid', 'name']):
        try:
            io_counters = proc.io_counters()
            app_name = proc.info['name']
            app_data[app_name][0] += io_counters.read_bytes / (1024 ** 2)  # Convert bytes to MB
            app_data[app_name][1] += io_counters.write_bytes / (1024 ** 2)  # Convert bytes to MB
        except (psutil.NoSuchProcess, psutil.AccessDenied, AttributeError):
            continue

    return app_data

def update_excel():
    while True:
        app_data = get_network_usage()

        # Clear existing data in the worksheet
        ws.delete_rows(2, ws.max_row)

        # Update Excel sheet with new data
        for app, usage in app_data.items():
            total_usage = usage[0] + usage[1]
            ws.append([app, total_usage])

        # Create a pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "App Data Usage"

        # Highlight the largest slice
        slices = [DataPoint(idx=i) for i in range(len(app_data))]
        slices[0].graphicalProperties.solidFill = "FF0000"  # Red
        pie.series[0].data_points = slices

        # Remove any existing chart
        if "Chart" in wb.sheetnames:
            wb.remove(wb["Chart"])

        chart_ws = wb.create_sheet("Chart")
        chart_ws.add_chart(pie, "A1")

        # Save the workbook
        wb.save(workbook_name)

        # Wait for 10 seconds before updating again
        time.sleep(10)

if __name__ == "__main__":
    update_excel()
